"""
pipeline.py — Tax document processing pipeline
Handles: PDF extraction via AWS Bedrock, file renaming, Excel population, packaging
"""
import base64
import concurrent.futures
import json
import os
import shutil
import time
import zipfile
from pathlib import Path

import openpyxl

from bedrock_client import BedrockTaxExtractor

# ── 1040 reporting order ──────────────────────────────────────────────────────
FORM_ORDER = [
    "W-2",
    "1099-INT",
    "1099-DIV",
    "1099-B",
    "1099-NEC",
    "1099-R",
    "1099-G",
    "1099-SA",
    "SSA-1099",
    "1098",
    "1098-E",
    "1098-T",
    "5498",
    "5498-SA",
    "Consolidated 1099",
]

# ── Bedrock/Sonnet extraction prompt ───────────────────────────────────────────
EXTRACTION_PROMPT = """You are a tax document processor for an accounting firm.
Extract all data from this tax form PDF.

Respond with ONLY a valid JSON object — no markdown, no explanation, no code fences.

Required fields:
{
  "form_type": "W-2",
  "tax_year": "2025",
  "taxpayer_name": "Jane Smith",
  "spouse_name": null,
  "all_names": ["Jane Smith"],
  "payer_name": "Employer or Institution Name",
  "state": "AL",
  "filing_status": null,
  "charity_cash": null,
  "charity_noncash": null,
  "medical_expenses": null,
  "boxes": {
    "box_1": 52000.00,
    "box_2": 6240.00
  },
  "field_metadata": {
    "form_type": {
      "confidence": 98,
      "evidence": "Form W-2 Wage and Tax Statement",
      "page": 1,
      "notes": ""
    },
    "boxes.box_1": {
      "confidence": 96,
      "evidence": "Box 1 Wages, tips, other compensation 52000.00",
      "page": 1,
      "notes": ""
    }
  },
  "validation_flags": [],
  "notes": ""
}

Reviewer metadata rules:
- Keep the main data fields exactly as scalar values, arrays, objects, or null as shown above.
- Put field-level confidence/evidence in field_metadata using the same top-level field name, or boxes.[box_key] for box values.
- confidence must be an integer from 0 to 100.
- evidence must be a short snippet copied from the PDF or extracted text that supports the value.
- page is optional and should be the visible source page number when available.
- notes is optional and should describe ambiguity, not restate the value.
- If evidence is unavailable, use an empty string and lower confidence appropriately.

Filing status rules:
- If the document or any attached notes explicitly state "MFJ", "Married Filing Jointly", "Married" with two names, set filing_status="MFJ"
- If explicitly "Single", "HOH", "Head of Household", "MFS", set accordingly
- Otherwise leave filing_status as null
- Always populate all_names with every distinct adult name found anywhere in the document
- If a spouse name is visible, populate spouse_name

Supplemental fields (populate from client notes or any source — null if not found):
- charity_cash: total cash/check charitable donations
- charity_noncash: total non-cash charitable donations (clothing, goods, etc.)
- medical_expenses: total gross medical expenses paid out of pocket
- car_tags: total vehicle registration / car tag fees paid
- taxpayer_dob: taxpayer date of birth (MM/DD/YYYY format if found)
- spouse_dob: spouse date of birth (MM/DD/YYYY format if found)
- additional_interest: interest income reported in notes not covered by a separate 1099-INT (e.g. savings account, money market) — total dollar amount
- additional_interest_payer: institution name for the additional interest if mentioned

Box extraction rules by form type:

W-2:
  box_1=wages, box_2=fed_wh, box_3=ss_wages, box_4=ss_tax,
  box_5=med_wages, box_6=med_tax, box_12_code=code(D/E/etc), box_12_amount=amount,
  box_16=state_wages, box_17=state_wh, box_15=state_abbr

1099-INT: box_1=interest, box_4=fed_wh, box_8=tax_exempt_interest
1099-DIV: box_1a=total_div, box_1b=qualified_div, box_2a=cap_gains_dist, box_4=fed_wh, box_7=foreign_tax_paid
1099-R:   box_1=gross_dist, box_2a=taxable_amt, box_4=fed_wh, box_7=dist_code, box_2b_taxable_not_determined=true/false
1099-NEC: box_1=nec_compensation, box_4=fed_wh
1099-B:   box_1d=proceeds, box_1e=cost_basis, box_2=term(short/long), box_4=fed_wh,
          box_1d_st=short_term_proceeds_total, box_1e_st=short_term_basis_total,
          box_1d_lt=long_term_proceeds_total, box_1e_lt=long_term_basis_total
          (If the form has both short and long term sections, populate the _st and _lt fields with each section's totals.
           If only one term is present, also set box_1d and box_1e to that total and box_2 to short or long.)
1099-G:   box_1=unemployment, box_4=fed_wh, box_11=state_income_tax_refund
1099-SA:  box_1=distributions, box_5=account_type(HSA/Archer/MA)
SSA-1099: box_3=net_benefits, box_5=repaid, box_6=vtw_fed_wh
1098:     box_1=mortgage_interest, box_5=property_tax, box_2=outstanding_principal, box_7=address_of_property
1098-E:   box_1=student_loan_interest
1098-T:   box_1=tuition_paid, box_5=scholarships
5498:     box_1=ira_contributions, box_2=rollover, box_10=roth_contributions
5498-SA:  box_2=hsa_contributions, box_5=account_type

validation_flags — add a string for each issue found:
- If tax_year is not 2025: "TAX YEAR MISMATCH: document is [year], expected 2025"
- If W-2 box_4 / box_3 does not equal ~6.2% (allow ±$5): "SS tax math error: expected X got Y" — ONLY add this flag if the amounts are actually different. Do NOT add if expected equals got.
- If W-2 box_6 / box_5 does not equal ~1.45% (allow ±$5): "Medicare tax math error: expected X got Y" — ONLY add this flag if the amounts are actually different. Do NOT add if expected equals got.
- If W-2 box_12_code is D and box_12_amount > 23500: "401k deferral exceeds $23,500 limit"
- If 1099-R box_2a > box_1: "Taxable amount exceeds gross distribution"
- If state is TN and box_17 > 0: "TN client shows state withholding — should be $0"

Use null for any box not present on the form.
"""

# ── 2025 limits for validation ────────────────────────────────────────────────
SIMPLE_FORMS = ["W-2", "1099-INT", "1099-DIV", "1099-NEC", "1098", "1099-R"]
COMPLEX_FORMS = ["1099-B", "Consolidated 1099", "client_notes", "client_organizer"]

# ── 2025 limits for validation ────────────────────────────────────────────────
LIMITS_2025 = {
    "ss_wage_base": 176100,
    "ss_rate": 0.062,
    "med_rate": 0.0145,
    "k401_limit_under50": 23500,
    "k401_limit_50plus": 31000,
    "ira_limit_under50": 7000,
    "ira_limit_50plus": 8000,
    "hsa_self": 4300,
    "hsa_family": 8550,
    "std_ded_single": 15000,
    "std_ded_mfj": 30000,
    "std_ded_hoh": 22500,
}


class TaxPipeline:
    def __init__(self, api_key, template_1040, template_doublecheck,
                 output_folder, log_callback=None, aws_region=None,
                 aws_profile=None, bedrock_model_id=None, extractor=None):
        # api_key is a legacy constructor argument kept so existing tests and
        # callers do not break; Bedrock uses the standard AWS credential chain.
        self.extractor = extractor or BedrockTaxExtractor(
            region=aws_region,
            model_id=bedrock_model_id,
            aws_profile=aws_profile,
        )
        self.template_1040 = template_1040
        self.template_doublecheck = template_doublecheck
        self.output_folder = output_folder
        self.log = log_callback or print
        self.extracted = []

    # ── Public entry point ────────────────────────────────────────────────────
    def run(self, pdf_paths, last_name, first_name):
        pipeline_start = time.time()
        client_slug = f"{last_name}_{first_name}_2025" if first_name else f"{last_name}_2025"
        out_dir = Path(self.output_folder) / client_slug

        # ── Create folder structure ───────────────────────────────────────────
        sd_dir  = out_dir / "SD"               # Source documents (PDFs + notes)
        rev_dir = out_dir / "Review"           # Excel workbooks
        ret_dir = out_dir / "Return"           # Empty — for completed return
        sig_dir = out_dir / "Signature Pages"  # Empty — for signature pages
        for d in [out_dir, sd_dir, rev_dir, ret_dir, sig_dir]:
            d.mkdir(parents=True, exist_ok=True)

        # ── STEP 1: Classify & extract (parallel) ────────────────────────────
        step1_start = time.time()
        self.log(f"STEP 1 — Classifying {len(pdf_paths)} documents via AWS Bedrock/Sonnet... (parallel, 4 workers)")
        self.extracted = []

        with concurrent.futures.ThreadPoolExecutor(max_workers=4) as executor:
            future_to_path = {executor.submit(self._extract_pdf, p): p for p in pdf_paths}
            raw_results = {}
            for future in concurrent.futures.as_completed(future_to_path):
                pdf_path = future_to_path[future]
                try:
                    data, elapsed = future.result()
                except Exception as e:
                    data, elapsed = None, 0.0
                    self.log(f"    ❌ Unexpected error processing {Path(pdf_path).name}: {e}")
                raw_results[pdf_path] = (data, elapsed)

        # Log results in 1040 reporting order after all futures complete
        def _result_sort_key(item):
            _, (data, _elapsed) = item
            if not data:
                return 100
            ft = (data.get("form_type") or "").lower().strip()
            if any(x in ft for x in ("client", "organizer", "notes")):
                return 99
            return self._form_sort_key(data.get("form_type", ""))

        for pdf_path, (data, elapsed) in sorted(raw_results.items(), key=_result_sort_key):
            self.log(f"  Reading: {Path(pdf_path).name}")
            if data:
                # Normalize form type aliases (case-insensitive)
                ft = (data.get("form_type") or "").lower().strip()
                if any(x in ft for x in ("client", "organizer", "notes")):
                    data["form_type"] = "client_notes"
                    data["payer_name"] = last_name  # Use client name, not inferred payer

                # Strip false-positive W-2 math flags (expected == got)
                flags = data.get("validation_flags") or []
                data["validation_flags"] = [
                    f for f in flags
                    if not ("math error" in f.lower() and
                            len(set(f.split("expected ")[-1].split(" got "))) == 1)
                ]

                self.extracted.append({"path": pdf_path, "data": data})
                form = data.get("form_type", "?")
                year = data.get("tax_year", "?")
                payer = data.get("payer_name", "?")
                self.log(f"    → {form} | {payer} | Tax year {year}  ({elapsed:.1f}s)")
                for flag in data.get("validation_flags", []):
                    self.log(f"    ⚠  {flag}")
            else:
                self.log(f"    ❌ Could not extract data from {Path(pdf_path).name}")

        step1_elapsed = time.time() - step1_start
        self.log(f"STEP 1 complete — {len(self.extracted)} documents classified in {step1_elapsed:.1f}s")
        self.log("")

        if not self.extracted:
            raise RuntimeError("No documents could be extracted. Check that files are readable PDFs.")

        # ── HARD STOP: year mismatch ─────────────────────────────────────────
        mismatched = [
            item for item in self.extracted
            if item["data"].get("tax_year") not in ("2025", None)
        ]
        if mismatched:
            self.log("")
            self.log("❌ HARD STOP — Year mismatch detected. Generating rejection package...")
            self._write_rejection_package(out_dir, client_slug, mismatched)
            zip_path = self._zip(out_dir, client_slug)
            self.log(f"  Rejection package ready: {zip_path}")
            return

        # ── Determine filing status ──────────────────────────────────────────
        filing_status, fs_confidence, fs_reason = self._determine_filing_status()
        self.log(f"  Filing status: {filing_status} ({fs_confidence}) — {fs_reason}")

        # ── STEP 3: Sort in 1040 order ───────────────────────────────────────
        self.log("STEP 3 — Sorting in 1040 reporting order...")
        self.extracted.sort(key=lambda x: self._form_sort_key(x["data"].get("form_type", "")))

        # ── STEP 2: Rename and copy PDFs ─────────────────────────────────────
        self.log("STEP 2 — Renaming PDFs...")
        doc_log_lines = [
            "DOCUMENT LOG",
            f"Client: {last_name}, {first_name}",
            f"Tax Year: 2025",
            "=" * 55,
            "",
            "FILE RENAMES",
            "-" * 40,
        ]
        for item in self.extracted:
            data = item["data"]
            payer_raw = data.get("payer_name") or "Unknown"
            payer_clean = "".join(c for c in payer_raw if c.isalnum())[:20]
            form = data.get("form_type", "Unknown").replace(" ", "-")
            new_name = f"{form}_{payer_clean}_2025.pdf"
            dest = sd_dir / new_name
            shutil.copy2(item["path"], dest)
            orig = Path(item["path"]).name
            doc_log_lines.append(f"  {orig}")
            doc_log_lines.append(f"    → {new_name}")
            doc_log_lines.append("")
            item["renamed"] = new_name
            self.log(f"  {orig} → {new_name}")

        # ── STEP 4: Populate spreadsheets ────────────────────────────────────
        self.log("STEP 4 — Populating spreadsheets...")
        if self.template_1040 and Path(self.template_1040).exists():
            self._populate_1040(rev_dir, client_slug, last_name, first_name, filing_status)
        else:
            self.log("  ⚠  1040 template not configured — skipping")

        if self.template_doublecheck and Path(self.template_doublecheck).exists():
            self._populate_doublecheck(rev_dir, client_slug, filing_status)
        else:
            self.log("  ⚠  DoubleCheck template not configured — skipping")

        # ── Document log: validation flags ───────────────────────────────────
        all_flags = []
        for item in self.extracted:
            for flag in item["data"].get("validation_flags", []):
                all_flags.append(f"  [{item['renamed']}] {flag}")

        doc_log_lines += [
            "VALIDATION FLAGS",
            "-" * 40,
        ]
        if all_flags:
            doc_log_lines += all_flags
        else:
            doc_log_lines.append("  None — all documents passed validation")

        doc_log_lines += [
            "",
            "OPEN ITEMS",
            "-" * 40,
        ]

        taxpayer_dob_str = next((item["data"].get("taxpayer_dob") for item in self.extracted if item["data"].get("taxpayer_dob")), None)
        spouse_dob_str   = next((item["data"].get("spouse_dob")   for item in self.extracted if item["data"].get("spouse_dob")),   None)

        open_items = []
        if taxpayer_dob_str and spouse_dob_str:
            open_items.append(f"  Birthdates provided — taxpayer: {taxpayer_dob_str}, spouse: {spouse_dob_str}.")
        elif taxpayer_dob_str:
            open_items.append(f"  Taxpayer birthdate provided: {taxpayer_dob_str}. Spouse birthdate assumed age 40 — confirm with client.")
        else:
            open_items.append("  Birthdates assumed (age 40) — confirm with client.")
        open_items.append("  If filer or spouse is 65+, senior standard deduction add-on applies.")

        # Filing status
        if fs_confidence != "explicit":
            open_items.append(f"  FILING STATUS: {fs_reason}")
        else:
            open_items.append(f"  Filing status: {filing_status} — confirmed from documents.")

        # Foreign tax
        for item in self.extracted:
            ft_form = item["data"].get("form_type", "")
            boxes = item["data"].get("boxes") or {}
            if ft_form == "1099-DIV" and _f(boxes.get("box_7")) > 0:
                open_items.append(
                    f"  FOREIGN TAX CREDIT: ${_f(boxes.get('box_7')):.2f} foreign tax paid "
                    f"({item['data'].get('payer_name','')}) — review Form 1116 eligibility."
                )

        # Non-cash charity thresholds
        for item in self.extracted:
            noncash = _f(item["data"].get("charity_noncash"))
            if noncash > 5000:
                open_items.append(f"  NON-CASH CHARITY (${noncash:.2f}): Form 8283 required + qualified appraisal (>$5,000).")
            elif noncash > 500:
                open_items.append(f"  NON-CASH CHARITY (${noncash:.2f}): Form 8283 required (>$500).")

        doc_log_lines += open_items

        doc_log_lines += [
            "",
            "EXTRACTED VALUES SUMMARY",
            "-" * 40,
        ]
        for item in self.extracted:
            d = item["data"]
            doc_log_lines.append(f"  {item['renamed']}")
            for k, v in (d.get("boxes") or {}).items():
                if v is not None and v != 0:
                    doc_log_lines.append(f"    {k}: {v}")
            if d.get("notes"):
                doc_log_lines.append(f"    notes: {d['notes']}")
            doc_log_lines.append("")

        doc_log_lines += [
            "REVIEWER FIELD METADATA",
            "-" * 40,
        ]
        metadata_lines = []
        for item in self.extracted:
            metadata = item["data"].get("field_metadata") or {}
            for field_path, details in metadata.items():
                if not isinstance(details, dict):
                    continue
                confidence = details.get("confidence", "?")
                evidence = details.get("evidence") or ""
                page = details.get("page")
                page_text = f" p.{page}" if page else ""
                line = f"  [{item['renamed']}] {field_path}: confidence {confidence}{page_text}"
                if evidence:
                    line += f" — evidence: {evidence}"
                if details.get("notes"):
                    line += f" — notes: {details['notes']}"
                metadata_lines.append(line)
        if metadata_lines:
            doc_log_lines += metadata_lines
        else:
            doc_log_lines.append("  None returned")

        log_path = out_dir / "document_log.txt"
        log_path.write_text("\n".join(doc_log_lines), encoding="utf-8")
        self.log("  document_log.txt written")

        # ── STEP 5: Zip ───────────────────────────────────────────────────────
        self.log("STEP 5 — Packaging...")
        zip_path = self._zip(out_dir, client_slug)
        self.log(f"  Package ready: {zip_path}")
        total_elapsed = time.time() - pipeline_start
        self.log(f"✅ Pipeline complete — total time: {total_elapsed:.1f}s")

    # ── Private: parse DOB and determine senior status ────────────────────────
    def _parse_dob(self, dob_str):
        """Parse MM/DD/YYYY or similar DOB string. Returns (year, month, day) or None."""
        if not dob_str:
            return None
        import re
        # Try MM/DD/YYYY or M/D/YYYY
        m = re.search(r'(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})', str(dob_str))
        if m:
            return int(m.group(3)), int(m.group(1)), int(m.group(2))
        # Try YYYY-MM-DD
        m = re.search(r'(\d{4})[/\-](\d{1,2})[/\-](\d{1,2})', str(dob_str))
        if m:
            return int(m.group(1)), int(m.group(2)), int(m.group(3))
        return None

    def _is_senior(self, dob_tuple):
        """Returns True if person turns 65 or older during tax year 2025."""
        if not dob_tuple:
            return False
        year, month, day = dob_tuple
        return (2025 - year) >= 65

    # ── Private: filing status determination ─────────────────────────────────
    def _determine_filing_status(self):
        """
        3-tier filing status logic:
        1. Explicit statement in any document → use it (highest confidence)
        2. Two distinct adult names across docs, no explicit status → likely MFJ, flag to confirm
        3. One name only → default Single, flag to confirm
        Never silently default — always return (status, confidence, reason).
        """
        # Tier 1: explicit filing status from any document
        for item in self.extracted:
            fs = item["data"].get("filing_status")
            if fs:
                return fs.upper(), "explicit", f"Filing status '{fs}' stated explicitly in {item['data'].get('form_type', 'document')}"

        # Tier 2: collect all distinct names across all documents
        all_names = set()
        for item in self.extracted:
            for name in (item["data"].get("all_names") or []):
                if name and str(name).strip():
                    all_names.add(str(name).strip().upper())
            tp = item["data"].get("taxpayer_name")
            if tp:
                all_names.add(str(tp).strip().upper())
            sp = item["data"].get("spouse_name")
            if sp:
                all_names.add(str(sp).strip().upper())

        if len(all_names) >= 2:
            names_display = ", ".join(sorted(all_names))
            return "MFJ", "inferred", f"Two distinct names found across documents ({names_display}) — assumed MFJ, confirm with client"

        # Tier 3: single name only
        name_display = next(iter(all_names), "unknown")
        return "Single", "default", f"Only one name found ({name_display}) — defaulted to Single, confirm with client"

    # ── Private: text layer extraction ───────────────────────────────────────
    def _extract_text_layer(self, pdf_path) -> "str | None":
        """
        Try to extract text from a digital PDF using pdfplumber.
        Returns extracted text string if meaningful text found (>200 chars),
        or None if scanned/image-only PDF.
        """
        try:
            import pdfplumber

            with pdfplumber.open(pdf_path) as pdf:
                text_parts = []
                for page in pdf.pages:
                    t = page.extract_text()
                    if t:
                        text_parts.append(t)
                full_text = "\n".join(text_parts).strip()
                return full_text if len(full_text) > 200 else None
        except Exception:
            return None

    def _select_model(self, form_type: str, has_text_layer: bool) -> str:
        """
        Sonnet is the only enabled extraction model for the Bedrock v1 epic.
        """
        return self.extractor.model_id

    # ── Private: Bedrock extraction ──────────────────────────────────────────
    def _extract_pdf(self, pdf_path):
        # Try text layer first — faster and fewer tokens for digital PDFs
        text_layer = self._extract_text_layer(pdf_path)

        if text_layer:
            content = [{
                "type": "text",
                "text": f"The following is extracted text from a tax document PDF:\n\n{text_layer}\n\n{EXTRACTION_PROMPT}"
            }]
        else:
            with open(pdf_path, "rb") as f:
                pdf_b64 = base64.standard_b64encode(f.read()).decode("utf-8")
            content = [
                {
                    "type": "document",
                    "source": {
                        "type": "base64",
                        "media_type": "application/pdf",
                        "data": pdf_b64,
                    }
                },
                {"type": "text", "text": EXTRACTION_PROMPT}
            ]

        model = self._select_model("", has_text_layer=text_layer is not None)
        t0 = time.time()
        try:
            text = self.extractor.extract_text(content, max_tokens=1500)
            elapsed = time.time() - t0
            # Strip markdown code fences if model adds them
            text = text.replace("```json", "").replace("```", "").strip()
            # Use raw_decode so extra content after the first JSON object
            # (e.g. two objects returned back-to-back) doesn't cause a failure
            obj, _ = json.JSONDecoder().raw_decode(text)
            return self._normalize_extraction_schema(obj), elapsed
        except json.JSONDecodeError as e:
            self.log(f"    ❌ JSON parse error: {e}")
            return None, time.time() - t0
        except Exception as e:
            self.log(f"    ❌ Bedrock error using {model}: {e}")
            return None, time.time() - t0

    def _normalize_extraction_schema(self, obj):
        """
        Preserve the existing flat extraction values while accepting reviewer
        wrappers if the model accidentally returns them inline.
        """
        if not isinstance(obj, dict):
            return obj

        normalized = dict(obj)
        metadata = normalized.get("field_metadata")
        if not isinstance(metadata, dict):
            metadata = {}

        wrapper_keys = {"value", "confidence", "evidence", "page", "notes"}

        def unwrap(field_path, current):
            if isinstance(current, dict) and "value" in current:
                if wrapper_keys.intersection(current.keys()):
                    field_meta = {
                        k: current.get(k)
                        for k in ("confidence", "evidence", "page", "notes")
                        if k in current
                    }
                    if field_meta:
                        metadata.setdefault(field_path, field_meta)
                    return current.get("value")
            return current

        for key in list(normalized.keys()):
            if key in ("boxes", "field_metadata"):
                continue
            normalized[key] = unwrap(key, normalized[key])

        boxes = normalized.get("boxes")
        if isinstance(boxes, dict):
            normalized_boxes = {}
            for box_key, box_value in boxes.items():
                normalized_boxes[box_key] = unwrap(f"boxes.{box_key}", box_value)
            normalized["boxes"] = normalized_boxes

        normalized["field_metadata"] = self._sanitize_field_metadata(metadata)
        return normalized

    def _sanitize_field_metadata(self, metadata):
        cleaned = {}
        for field_path, details in (metadata or {}).items():
            if not isinstance(details, dict):
                continue
            item = {}
            confidence = details.get("confidence")
            if confidence is not None:
                try:
                    item["confidence"] = max(0, min(100, int(confidence)))
                except (TypeError, ValueError):
                    pass
            if "evidence" in details:
                item["evidence"] = "" if details.get("evidence") is None else str(details.get("evidence"))[:500]
            if "page" in details:
                item["page"] = details.get("page")
            if "notes" in details:
                item["notes"] = "" if details.get("notes") is None else str(details.get("notes"))[:500]
            if item:
                cleaned[str(field_path)] = item
        return cleaned

    # ── Private: Excel — Glenn Reeves 1040 template ──────────────────────────
    def _populate_1040(self, out_dir, client_slug, last_name, first_name, filing_status="Single"):
        import tempfile, os
        # Read as raw bytes then write to temp — bypasses Mac quarantine xattr
        tmp_path = tempfile.mktemp(suffix='.xlsx')
        with open(self.template_1040, 'rb') as f:
            raw = f.read()
        with open(tmp_path, 'wb') as f:
            f.write(raw)
        wb = openpyxl.load_workbook(tmp_path)
        ws_1040 = wb["1040"] if "1040" in wb.sheetnames else wb.active

        # Taxpayer name
        ws_1040["B11"] = f"{last_name}, {first_name}".strip(", ")

        # ── Filing status checkboxes ──────────────────────────────────────────
        # Clear all first, then set the correct one
        # File_Single=F20, File_Marr_Joint=F22, MFS=F24, HOH=AB20, QSS=AB22
        fs_cells = {"F20": "", "F22": "", "F24": "", "AB20": "", "AB22": ""}
        fs_lookup = {
            "SINGLE": "F20",
            "MFJ":    "F22",
            "MFS":    "F24",
            "HOH":    "AB20",
            "QSS":    "AB22",
        }
        target_cell = fs_lookup.get(filing_status.upper(), "F22")  # Default MFJ
        for cell in fs_cells:
            ws_1040[cell] = "X" if cell == target_cell else ""

        # ── Birthdates — use actual if provided, default age 40 ──────────────
        # Collect DOBs from all extracted documents
        taxpayer_dob_str = next((item["data"].get("taxpayer_dob") for item in self.extracted if item["data"].get("taxpayer_dob")), None)
        spouse_dob_str   = next((item["data"].get("spouse_dob")   for item in self.extracted if item["data"].get("spouse_dob")),   None)

        tp_dob = self._parse_dob(taxpayer_dob_str)
        sp_dob = self._parse_dob(spouse_dob_str)

        # Taxpayer birthdate
        if tp_dob:
            ws_1040["AX15"] = tp_dob[0]; ws_1040["AY15"] = tp_dob[1]; ws_1040["AZ15"] = tp_dob[2]
        else:
            ws_1040["AX15"] = 1985; ws_1040["AY15"] = 1; ws_1040["AZ15"] = 1  # Age 40 default

        # Spouse birthdate
        if sp_dob:
            ws_1040["AX17"] = sp_dob[0]; ws_1040["AY17"] = sp_dob[1]; ws_1040["AZ17"] = sp_dob[2]
        else:
            ws_1040["AX17"] = 1985; ws_1040["AY17"] = 1; ws_1040["AZ17"] = 1  # Age 40 default

        # Glenn Reeves template: column D = Payer/Employer #1, E = #2, etc.
        # Track next available column per sheet
        next_col = {}

        for item in self.extracted:
            data = item["data"]
            ft = data.get("form_type", "")
            boxes = data.get("boxes") or {}
            sheet_name = self._sheet_for(wb, ft)
            # ── Forms that don't use column-rotation sheets ───────────────────
            if ft == "1099-B":
                if "Sch. D" in wb.sheetnames:
                    ws_schd = wb["Sch. D"]
                    st_proceeds = _f(boxes.get("box_1d_st")) or (_f(boxes.get("box_1d")) if str(boxes.get("box_2","")).lower() == "short" else 0)
                    st_basis    = _f(boxes.get("box_1e_st")) or (_f(boxes.get("box_1e")) if str(boxes.get("box_2","")).lower() == "short" else 0)
                    lt_proceeds = _f(boxes.get("box_1d_lt")) or (_f(boxes.get("box_1d")) if str(boxes.get("box_2","")).lower() == "long" else 0)
                    lt_basis    = _f(boxes.get("box_1e_lt")) or (_f(boxes.get("box_1e")) if str(boxes.get("box_2","")).lower() == "long" else 0)
                    if st_proceeds: ws_schd["I17"] = st_proceeds   # Line 1a proceeds
                    if st_basis:    ws_schd["K17"] = st_basis      # Line 1a basis
                    if lt_proceeds: ws_schd["I35"] = lt_proceeds   # Line 8a proceeds
                    if lt_basis:    ws_schd["K35"] = lt_basis      # Line 8a basis
                continue

            if ft == "client_notes":
                # Write any interest income from notes (savings accounts, etc.)
                # to the next available column in the 1099-INT sheet.
                # Skip if payer already has a 1099-INT in the packet (avoid double-count).
                add_interest = _f(data.get("additional_interest"))
                add_payer = data.get("additional_interest_payer")
                if add_interest and not self._interest_already_covered(add_payer):
                    int_sheet = self._sheet_for(wb, "1099-INT")
                    if int_sheet and int_sheet in wb.sheetnames:
                        ws_int = wb[int_sheet]
                        col = next_col.get(int_sheet, 4)
                        ws_int.cell(row=6, column=col).value = add_interest
                        next_col[int_sheet] = col + 1
                continue

            if not sheet_name or sheet_name not in wb.sheetnames:
                continue

            ws = wb[sheet_name]
            col = next_col.get(sheet_name, 4)  # Column D = 4

            if ft == "W-2":
                ws.cell(row=3, column=col).value = _f(boxes.get("box_1"))   # Wages
                ws.cell(row=4, column=col).value = _f(boxes.get("box_2"))   # Fed WH
                ws.cell(row=5, column=col).value = _f(boxes.get("box_3"))   # SS wages
                ws.cell(row=6, column=col).value = _f(boxes.get("box_4"))   # SS tax
                ws.cell(row=7, column=col).value = _f(boxes.get("box_5"))   # Med wages
                ws.cell(row=8, column=col).value = _f(boxes.get("box_6"))   # Med tax

            elif ft == "1099-INT":
                ws.cell(row=6, column=col).value = _f(boxes.get("box_1"))   # Interest
                ws.cell(row=10, column=col).value = _f(boxes.get("box_4"))  # Fed WH

            elif ft == "1099-DIV":
                ws.cell(row=6, column=col).value = _f(boxes.get("box_1a"))  # Total divs
                ws.cell(row=7, column=col).value = _f(boxes.get("box_1b"))  # Qualified
                ws.cell(row=8, column=col).value = _f(boxes.get("box_2a"))  # Cap gains
                ws.cell(row=13, column=col).value = _f(boxes.get("box_4"))  # Fed WH
                ws.cell(row=16, column=col).value = _f(boxes.get("box_7"))  # Foreign tax

            elif ft == "1099-NEC":
                ws.cell(row=6, column=col).value = _f(boxes.get("box_1"))   # NEC comp
                ws.cell(row=9, column=col).value = _f(boxes.get("box_4"))   # Fed WH

            elif ft == "1099-R":
                dist_code = str(boxes.get("box_7") or "").strip()
                notes_upper = (data.get("notes") or "").upper()
                is_ira = dist_code in ("1", "2", "3", "4") or "IRA" in notes_upper
                ws.cell(row=3, column=col).value = "IRA" if is_ira else "P/A"
                ws.cell(row=4, column=col).value = _f(boxes.get("box_1"))   # Gross
                ws.cell(row=5, column=col).value = _f(boxes.get("box_2a"))  # Taxable
                ws.cell(row=8, column=col).value = _f(boxes.get("box_4"))   # Fed WH

            elif ft == "1099-G":
                ws.cell(row=6, column=col).value = _f(boxes.get("box_1"))   # Unemployment
                ws.cell(row=9, column=col).value = _f(boxes.get("box_4"))   # Fed WH

            elif ft == "SSA-1099":
                # Fixed layout: taxpayer rows 4/7, spouse rows 9/12 — no column rotation
                ws.cell(row=4, column=4).value = _f(boxes.get("box_3"))     # Benefits paid
                ws.cell(row=7, column=4).value = _f(boxes.get("box_6"))     # Fed WH
                continue  # Skip column increment for SSA-1099

            next_col[sheet_name] = col + 1

        # ── Aggregate supplemental fields across all documents ────────────────
        total_medical = sum(_f(item["data"].get("medical_expenses")) for item in self.extracted)
        total_charity_cash = sum(_f(item["data"].get("charity_cash")) for item in self.extracted)
        total_charity_noncash = sum(_f(item["data"].get("charity_noncash")) for item in self.extracted)
        total_foreign_tax = sum(
            _f((item["data"].get("boxes") or {}).get("box_7"))
            for item in self.extracted
            if item["data"].get("form_type") == "1099-DIV"
        )

        # Write to Sch. A if present
        if "Sch. A" in wb.sheetnames:
            ws_scha = wb["Sch. A"]
            if total_medical:
                ws_scha["R16"] = total_medical
            if total_charity_cash:
                ws_scha["N52"] = total_charity_cash
            if total_charity_noncash and total_charity_noncash <= 500:
                ws_scha["N54"] = total_charity_noncash

        out_path = out_dir / f"{client_slug}_1040.xlsx"
        wb.save(str(out_path))
        try:
            os.unlink(tmp_path)
        except Exception:
            pass
        self.log(f"  1040 saved: {out_path.name}")

    # ── Private: Excel — DoubleCheck template ─────────────────────────────────
    def _populate_doublecheck(self, out_dir, client_slug, filing_status="Single"):
        import tempfile, os
        tmp_path = tempfile.mktemp(suffix='.xlsx')
        with open(self.template_doublecheck, 'rb') as f:
            raw = f.read()
        with open(tmp_path, 'wb') as f:
            f.write(raw)
        wb = openpyxl.load_workbook(tmp_path)
        ws = wb.active

        # ── Filing status — check for senior status ───────────────────────────
        taxpayer_dob_str = next((item["data"].get("taxpayer_dob") for item in self.extracted if item["data"].get("taxpayer_dob")), None)
        spouse_dob_str   = next((item["data"].get("spouse_dob")   for item in self.extracted if item["data"].get("spouse_dob")),   None)
        tp_senior = self._is_senior(self._parse_dob(taxpayer_dob_str))
        sp_senior = self._is_senior(self._parse_dob(spouse_dob_str))
        either_senior = tp_senior or sp_senior

        # Map filing status to correct checkbox (regular vs senior)
        # H2=Single, H3=MFJ, H4=HOH, H5=MFS, H6=SENIOR Single, H7=SENIOR MFJ, H8=SENIOR HOH
        if either_senior:
            fs_map = {"SINGLE": "H6", "MFJ": "H7", "HOH": "H8", "MFS": "H5"}
        else:
            fs_map = {"SINGLE": "H2", "MFJ": "H3", "HOH": "H4", "MFS": "H5"}

        for row in range(2, 9):
            ws.cell(row=row, column=8).value = False
        target = fs_map.get(filing_status.upper(), "H3")
        ws[target] = True

        # Determine if TN client (all docs from TN → no state WH)
        states = [item["data"].get("state", "") for item in self.extracted]
        is_tn = all(s.upper() == "TN" for s in states if s)

        # Track next W-2 data entry row (rows 3-8, one per W-2)
        next_w2_row = 3

        for item in self.extracted:
            data = item["data"]
            ft = data.get("form_type", "")
            boxes = data.get("boxes") or {}

            if ft == "W-2":
                if next_w2_row <= 8:   # Max 6 W-2s (rows 3-8); row 9 is SUM formula
                    ws.cell(row=next_w2_row, column=2).value = _f(boxes.get("box_1"))   # B — Fed wages
                    ws.cell(row=next_w2_row, column=3).value = _f(boxes.get("box_16"))  # C — AL wages
                    ws.cell(row=next_w2_row, column=4).value = _f(boxes.get("box_2"))   # D — Fed WH
                    ws.cell(row=next_w2_row, column=5).value = 0 if is_tn else _f(boxes.get("box_17"))  # E — AL WH
                    ws.cell(row=next_w2_row, column=6).value = _f(boxes.get("box_4")) + _f(boxes.get("box_6"))  # F — FICA
                    next_w2_row += 1

            elif ft == "1099-INT":
                ws["B11"] = _f(boxes.get("box_1"))  # Interest income
                ws["C11"] = _f(boxes.get("box_4"))  # Federal WH

            elif ft == "1099-DIV":
                # B9/C9 are =SUM(B3:B8) formula cells — NEVER write to them
                ws["B16"] = _f(boxes.get("box_1a"))   # Ordinary dividends
                if _f(boxes.get("box_4")) > 0:
                    ws["C16"] = _f(boxes.get("box_4"))  # Fed WH on dividends
                # Foreign tax paid
                if _f(boxes.get("box_7")) > 0:
                    ws["B60"] = _f(boxes.get("box_7"))
                # Cap gain distributions → SCH D sheet, E29 (data entry row)
                cap_gain = _f(boxes.get("box_2a"))
                if cap_gain:
                    if "SCH D" in wb.sheetnames:
                        wb["SCH D"]["E29"] = cap_gain

            elif ft == "1099-NEC":
                ws["G3"] = _f(boxes.get("box_1"))   # NEC income
                ws["C8"] = _f(boxes.get("box_4"))   # Federal WH
                if "SCH C" in wb.sheetnames:
                    wb["SCH C"]["B3"] = _f(boxes.get("box_1"))

            elif ft == "1099-B":
                if "SCH D" in wb.sheetnames:
                    ws_schd = wb["SCH D"]
                    # ST proceeds/basis → SCH D B3/C3
                    st_proceeds = _f(boxes.get("box_1d_st")) or (_f(boxes.get("box_1d")) if str(boxes.get("box_2","")).lower() == "short" else 0)
                    st_basis    = _f(boxes.get("box_1e_st")) or (_f(boxes.get("box_1e")) if str(boxes.get("box_2","")).lower() == "short" else 0)
                    # LT proceeds/basis → SCH D B18/C18
                    lt_proceeds = _f(boxes.get("box_1d_lt")) or (_f(boxes.get("box_1d")) if str(boxes.get("box_2","")).lower() == "long" else 0)
                    lt_basis    = _f(boxes.get("box_1e_lt")) or (_f(boxes.get("box_1e")) if str(boxes.get("box_2","")).lower() == "long" else 0)
                    if st_proceeds: ws_schd["B3"]  = st_proceeds
                    if st_basis:    ws_schd["C3"]  = st_basis
                    if lt_proceeds: ws_schd["B18"] = lt_proceeds
                    if lt_basis:    ws_schd["C18"] = lt_basis

            elif ft == "1099-R":
                dist_code = str(boxes.get("box_7") or "").strip()
                gross = _f(boxes.get("box_1"))
                fed_wh = _f(boxes.get("box_4"))
                notes_upper = (data.get("notes") or "").upper()
                is_ira = dist_code in ("1", "2", "3", "4") or "IRA" in notes_upper
                if is_ira:
                    ws["B20"] = gross
                    ws["C20"] = fed_wh
                else:
                    ws["B22"] = gross
                    ws["C22"] = fed_wh

            elif ft == "1099-G":
                ws["B10"] = _f(boxes.get("box_1"))  # Unemployment
                ws["C10"] = _f(boxes.get("box_4"))

            elif ft == "SSA-1099":
                ws["B23"] = _f(boxes.get("box_3"))  # Social Security benefits
                ws["C23"] = _f(boxes.get("box_6"))

            elif ft == "1098":
                ws["K25"] = _f(boxes.get("box_1"))  # Mortgage interest
                if _f(boxes.get("box_5")) > 0:
                    ws["K22"] = _f(boxes.get("box_5"))  # Property taxes

        # ── Interest from client notes (savings accounts not on a 1099-INT) ───
        # Skipped if payer matches an existing 1099-INT document (avoid double-count).
        for item in self.extracted:
            if item["data"].get("form_type") == "client_notes":
                add_interest = _f(item["data"].get("additional_interest"))
                add_payer = item["data"].get("additional_interest_payer")
                if add_interest and not self._interest_already_covered(add_payer):
                    ws["B11"] = _f(ws["B11"].value) + add_interest

        # ── Aggregate supplemental fields ─────────────────────────────────────
        total_medical = sum(_f(item["data"].get("medical_expenses")) for item in self.extracted)
        total_charity_cash = sum(_f(item["data"].get("charity_cash")) for item in self.extracted)
        total_charity_noncash = sum(_f(item["data"].get("charity_noncash")) for item in self.extracted)
        total_car_tags = sum(_f(item["data"].get("car_tags")) for item in self.extracted)

        if total_medical:
            ws["O21"] = total_medical
        if total_charity_cash:
            ws["K26"] = total_charity_cash
        if total_charity_noncash <= 500:
            ws["K27"] = total_charity_noncash  # ≤$500: populate directly
        # >$500 non-cash: logged as Open Item only — do not populate (Form 8283 required)
        if total_car_tags:
            ws["K23"] = total_car_tags

        out_path = out_dir / f"{client_slug}_DoubleCheck.xlsx"
        wb.save(str(out_path))
        try:
            os.unlink(tmp_path)
        except Exception:
            pass
        self.log(f"  DoubleCheck saved: {out_path.name}")

    # ── Private: rejection package ────────────────────────────────────────────
    def _write_rejection_package(self, out_dir, client_slug, mismatched):
        lines = [
            "REJECTION LOG",
            "=" * 55,
            "",
            "REASON: Tax year mismatch — processing halted",
            "Expected tax year: 2025",
            "",
            "MISMATCHED DOCUMENTS:",
        ]
        for item in mismatched:
            yr = item["data"].get("tax_year", "unknown")
            lines.append(f"  {Path(item['path']).name}  →  tax year {yr}")

        # Carry forward non-year validation flags
        other_flags = []
        for item in self.extracted:
            for flag in item["data"].get("validation_flags", []):
                if "year" not in flag.lower() and "mismatch" not in flag.lower():
                    other_flags.append(f"  [{Path(item['path']).name}] {flag}")
        if other_flags:
            lines += ["", "ADDITIONAL FINDINGS (carry forward to resubmission):"]
            lines += other_flags

        (out_dir / "rejection_log.txt").write_text("\n".join(lines), encoding="utf-8")

        client_last = client_slug.split("_")[0]
        req_lines = [
            f"Dear {client_last} Family,",
            "",
            "Thank you for submitting your tax documents. After reviewing your packet,",
            "we found that one or more documents are from a prior tax year.",
            "",
            "To complete your 2025 tax return, please provide current-year (2025)",
            "versions of the following documents:",
            "",
        ]
        for item in mismatched:
            d = item["data"]
            req_lines.append(f"  • {d.get('form_type', 'Document')} from {d.get('payer_name', 'your issuer')}")
        req_lines += [
            "",
            "Please resubmit your complete packet once you have received all 2025 documents.",
            "Most issuers mail tax documents by January 31, 2026.",
            "",
            "If you believe this is an error, please contact our office.",
            "",
            "Thank you,",
            "[Firm Name]",
            "Pelham, AL",
        ]
        (out_dir / "client_request.txt").write_text("\n".join(req_lines), encoding="utf-8")

    # ── Private: helpers ──────────────────────────────────────────────────────
    def _interest_already_covered(self, payer_name) -> bool:
        """
        Returns True if payer_name loosely matches an already-extracted 1099-INT
        document, meaning the interest is already accounted for and should not be
        double-counted from client notes.
        Normalizes both names to alphanumeric lowercase for comparison.
        """
        if not payer_name:
            return False
        norm = lambda s: "".join(c.lower() for c in str(s) if c.isalnum())
        p = norm(payer_name)
        for item in self.extracted:
            if item["data"].get("form_type") == "1099-INT":
                existing = norm(item["data"].get("payer_name") or "")
                if existing and (p in existing or existing in p):
                    return True
        return False

    def _form_sort_key(self, form_type):
        try:
            return FORM_ORDER.index(form_type)
        except ValueError:
            return 99

    def _sheet_for(self, wb, form_type):
        """Find the matching sheet name in the workbook for a given form type."""
        candidates = {
            "W-2":      ["W-2s", "W-2", "W2", "Wages"],
            "1099-INT": ["1099-INT", "1099INT", "Interest"],
            "1099-DIV": ["1099-DIV", "1099DIV", "Dividends"],
            "1099-R":   ["1099-R", "1099R", "Retirement"],
            "1099-NEC": ["1099-NEC", "1099NEC", "NEC"],
            "1099-G":   ["1099-G", "1099G"],
            "SSA-1099": ["SSA-1099", "SSA1099"],
            "1098":     ["1098", "Mortgage", "Deductions", "Sch. A"],
        }
        for name in candidates.get(form_type, [form_type]):
            if name in wb.sheetnames:
                return name
        return None

    def _write_to_next_row(self, ws, col_value_map, start_row=8):
        """Write values to the next empty row in a sheet starting at start_row."""
        row = start_row
        while ws.cell(row=row, column=2).value not in (None, "", 0):
            row += 1
            if row > start_row + 50:
                break
        for col_letter, value in col_value_map.items():
            if value is not None:
                ws[f"{col_letter}{row}"] = value

    def _zip(self, out_dir, client_slug):
        zip_path = Path(self.output_folder) / f"{client_slug}.zip"
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for f in out_dir.rglob("*"):
                if f.is_file():
                    zf.write(f, f.relative_to(Path(self.output_folder)))
        return zip_path


def _f(val):
    """Return float or 0 — never None into Excel."""
    if val is None:
        return 0
    try:
        return float(val)
    except (TypeError, ValueError):
        return 0
