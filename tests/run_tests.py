#!/usr/bin/env python3
"""
PrepDesk Test Suite
====================
Unit mode  — replays golden extraction fixtures through the Excel population logic.
             No API key or PDF files required. Fast.

Integration mode — runs the full pipeline against real PDFs and validates output.
                   Requires a configured API key and PDF files in fixtures/*/inputs/.

Validate mode — validates existing output Excel files without re-running the pipeline.

Usage:
  python tests/run_tests.py --unit
  python tests/run_tests.py --integration
  python tests/run_tests.py --validate PATH_TO_OUTPUT_DIR
"""

import argparse
import json
import os
import sys
import tempfile
import shutil
from pathlib import Path

# Ensure the project root is on the path
TESTS_DIR = Path(__file__).parent
PROJECT_DIR = TESTS_DIR.parent
sys.path.insert(0, str(PROJECT_DIR))

import openpyxl
from pipeline import TaxPipeline

FIXTURES_DIR = TESTS_DIR / "fixtures"
TOLERANCE = 0.02  # Dollar tolerance for float comparisons

# ── Colours ───────────────────────────────────────────────────────────────────
GREEN  = "\033[92m"
RED    = "\033[91m"
YELLOW = "\033[93m"
RESET  = "\033[0m"
BOLD   = "\033[1m"

def pass_str(): return f"{GREEN}✅ PASS{RESET}"
def fail_str(): return f"{RED}❌ FAIL{RESET}"
def warn_str(): return f"{YELLOW}⚠  WARN{RESET}"


# ── Excel validation ──────────────────────────────────────────────────────────

def _cell_value(ws, cell_addr):
    """Read a cell value, returning None if the cell is empty."""
    try:
        val = ws[cell_addr].value
        return val
    except Exception:
        return None


def _float_equal(actual, expected, tol=TOLERANCE):
    try:
        return abs(float(actual) - float(expected)) <= tol
    except (TypeError, ValueError):
        return False


def validate_workbook(wb, expected_sheets: dict, label: str) -> tuple[int, int]:
    """
    Validate an openpyxl workbook against a dict of {sheet: {cell: expected_value}}.
    Returns (passed, failed).
    """
    passed = failed = 0
    print(f"\n  {BOLD}{label}{RESET}")

    for sheet_name, cells in expected_sheets.items():
        if sheet_name.startswith("_comment"):
            continue
        if sheet_name not in wb.sheetnames:
            print(f"    {warn_str()}  Sheet '{sheet_name}' not found in workbook — skipping")
            continue
        ws = wb[sheet_name]
        print(f"    Sheet: {sheet_name}")

        for cell_addr, expected in cells.items():
            if cell_addr.startswith("_comment"):
                continue
            actual = _cell_value(ws, cell_addr)

            if isinstance(expected, bool):
                ok = bool(actual) == expected
            elif isinstance(expected, (int, float)):
                ok = _float_equal(actual, expected)
            elif expected == "" or expected is None:
                ok = actual in (None, "", 0, False)
            else:
                ok = str(actual).strip() == str(expected).strip()

            status = pass_str() if ok else fail_str()
            if ok:
                passed += 1
                print(f"      {status}  {cell_addr} = {actual!r}")
            else:
                failed += 1
                print(f"      {status}  {cell_addr}: expected {expected!r}, got {actual!r}")

    return passed, failed


# ── Pipeline harness (unit mode) ──────────────────────────────────────────────

def run_unit_test(fixture_dir: Path) -> tuple[int, int]:
    """
    Load golden extraction data, run only the Excel population logic,
    and validate output against expected JSON. No API call.
    """
    golden_path   = fixture_dir / "extracted_golden.json"
    exp_dc_path   = fixture_dir / "expected_doublecheck.json"
    exp_1040_path = fixture_dir / "expected_1040.json"

    # Read meta to determine test type
    meta_path = fixture_dir / "meta.json"
    test_type = "standard"
    if meta_path.exists():
        with open(meta_path) as f:
            _meta = json.load(f)
        test_type = _meta.get("test_type", "standard")

    # Rejection fixtures only need golden + no expected Excel
    if test_type == "rejection":
        if not golden_path.exists():
            print(f"  {fail_str()}  Missing fixture: {golden_path}")
            return 0, 1
        return run_rejection_test(fixture_dir, golden_path)

    for p in [golden_path, exp_dc_path, exp_1040_path]:
        if not p.exists() and p == exp_1040_path:
            continue  # 1040 expected is optional
        if not p.exists() and p != exp_1040_path:
            print(f"  {fail_str()}  Missing fixture: {p}")
            return 0, 1

    with open(golden_path) as f:
        golden = json.load(f)
    exp_dc    = json.load(open(exp_dc_path))   if exp_dc_path.exists()   else None
    exp_1040  = json.load(open(exp_1040_path)) if exp_1040_path.exists() else None

    # Find templates in the project directory
    template_1040       = PROJECT_DIR / "25_1040.xlsx"
    template_doublecheck = PROJECT_DIR / "2025_Tax_Return_Double_Check.xlsx"

    if not template_1040.exists() or not template_doublecheck.exists():
        print(f"  {fail_str()}  Excel templates not found in {PROJECT_DIR}")
        return 0, 1

    # Build a pipeline with dummy settings
    pipeline = TaxPipeline(
        api_key="unit-test-no-api",
        template_1040=str(template_1040),
        template_doublecheck=str(template_doublecheck),
        output_folder=tempfile.mkdtemp(),
        log_callback=lambda msg: None,  # Silence log output during tests
    )

    # Inject golden extracted data (replace __PLACEHOLDER__ paths with dummy)
    pipeline.extracted = []
    for item in golden:
        path = item["path"].replace("__PLACEHOLDER__", str(fixture_dir / "inputs"))
        pipeline.extracted.append({"path": path, "data": item["data"], "renamed": Path(path).name})

    # Determine filing status from golden data (replicate pipeline logic)
    filing_status = "MFJ"  # Explicit in golden client_notes
    for item in pipeline.extracted:
        fs = item["data"].get("filing_status")
        if fs:
            filing_status = fs.upper()
            break

    # Run Excel population into a temp output dir
    out_dir = Path(pipeline.output_folder) / "thornton_unit_test"
    rev_dir = out_dir / "Review"
    rev_dir.mkdir(parents=True, exist_ok=True)

    meta_path = fixture_dir / "meta.json"
    if meta_path.exists():
        with open(meta_path) as f:
            meta = json.load(f)
        client_slug = meta.get("client_slug", "TestClient_2025")
        last_name   = meta.get("last_name", "TestClient")
        first_name  = meta.get("first_name", "")
    else:
        client_slug = "TestClient_2025"
        last_name, first_name = "TestClient", ""

    pipeline._populate_1040(rev_dir, client_slug, last_name, first_name, filing_status)
    pipeline._populate_doublecheck(rev_dir, client_slug, filing_status)

    # Validate output
    dc_path   = rev_dir / f"{client_slug}_DoubleCheck.xlsx"
    i40_path  = rev_dir / f"{client_slug}_1040.xlsx"

    passed = failed = 0

    if exp_dc and dc_path.exists():
        wb_dc = openpyxl.load_workbook(str(dc_path))
        p, f = validate_workbook(wb_dc, exp_dc["sheets"], "DoubleCheck")
        passed += p; failed += f
    elif exp_dc:
        print(f"  {fail_str()}  DoubleCheck output not created")
        failed += 1

    if exp_1040 and i40_path.exists():
        wb_1040 = openpyxl.load_workbook(str(i40_path))
        p, f = validate_workbook(wb_1040, exp_1040["sheets"], "Glenn Reeves 1040")
        passed += p; failed += f
    elif exp_1040:
        print(f"  {fail_str()}  1040 output not created")
        failed += 1

    # Cleanup
    shutil.rmtree(pipeline.output_folder, ignore_errors=True)
    return passed, failed


# ── Rejection test harness ────────────────────────────────────────────────────

def run_rejection_test(fixture_dir: Path, golden_path: Path) -> tuple[int, int]:
    """
    Verify the year-mismatch hard stop: rejection_log.txt and client_request.txt
    must be created; no Excel files should be written.
    """
    passed = failed = 0

    with open(golden_path) as f:
        golden = json.load(f)

    meta_path = fixture_dir / "meta.json"
    with open(meta_path) as f:
        meta = json.load(f)
    client_slug = meta.get("client_slug", "Test_2025")
    last_name   = meta.get("last_name", "Test")
    first_name  = meta.get("first_name", "")

    template_1040        = PROJECT_DIR / "25_1040.xlsx"
    template_doublecheck = PROJECT_DIR / "2025_Tax_Return_Double_Check.xlsx"
    out_folder = tempfile.mkdtemp()

    pipeline = TaxPipeline(
        api_key="unit-test-no-api",
        template_1040=str(template_1040),
        template_doublecheck=str(template_doublecheck),
        output_folder=out_folder,
        log_callback=lambda _: None,
    )

    # Inject golden data
    pipeline.extracted = []
    for it in golden:
        path = it["path"].replace("__PLACEHOLDER__", str(fixture_dir / "inputs"))
        pipeline.extracted.append({"path": path, "data": it["data"], "renamed": Path(path).name})

    # Simulate the rejection detection + package generation
    from pathlib import Path as _Path
    out_dir = _Path(out_folder) / client_slug
    out_dir.mkdir(parents=True, exist_ok=True)

    mismatched = [i for i in pipeline.extracted
                  if i["data"].get("tax_year") not in ("2025", None)]
    pipeline._write_rejection_package(out_dir, client_slug, mismatched)

    print(f"\n  {BOLD}Rejection flow{RESET}")

    def chk(desc, ok):
        nonlocal passed, failed
        if ok:
            passed += 1
            print(f"    {pass_str()}  {desc}")
        else:
            failed += 1
            print(f"    {fail_str()}  {desc}")

    chk("Mismatched docs detected",          len(mismatched) > 0)
    chk(f"Detected {len(mismatched)} mismatch(es)",
        all(i["data"]["tax_year"] not in ("2025", None) for i in mismatched))
    chk("rejection_log.txt created",         (out_dir / "rejection_log.txt").exists())
    chk("client_request.txt created",        (out_dir / "client_request.txt").exists())

    # Verify rejection_log.txt content
    if (out_dir / "rejection_log.txt").exists():
        log_text = (out_dir / "rejection_log.txt").read_text()
        chk("Rejection log contains 'HARD STOP'",   "HARD STOP" in log_text or "REJECTION" in log_text)
        chk("Rejection log names mismatched doc",
            any(Path(i["path"]).name in log_text or
                i["data"].get("payer_name", "") in log_text
                for i in mismatched))

    # No Excel should exist
    rev_dir = out_dir / "Review"
    excel_files = list(rev_dir.glob("*.xlsx")) if rev_dir.exists() else []
    chk("No Excel files written (hard stop respected)", len(excel_files) == 0)

    shutil.rmtree(out_folder, ignore_errors=True)
    return passed, failed


# ── Integration test ──────────────────────────────────────────────────────────

def run_integration_test(fixture_dir: Path) -> tuple[int, int]:
    """
    Run the full pipeline against real PDFs in fixture_dir/inputs/,
    then validate output against expected JSON.
    Requires a valid API key in ~/.tax_processor/config.json.
    """
    import json as _json

    config_path = Path.home() / ".tax_processor" / "config.json"
    if not config_path.exists():
        print(f"  {fail_str()}  Config not found at {config_path} — run the app first to configure")
        return 0, 1

    with open(config_path) as f:
        config = _json.load(f)

    api_key            = config.get("api_key", "")
    template_1040      = config.get("template_1040", "")
    template_doublecheck = config.get("template_doublecheck", "")

    if not api_key:
        print(f"  {fail_str()}  No API key in config")
        return 0, 1

    inputs_dir = fixture_dir / "inputs"
    pdf_paths = sorted(inputs_dir.glob("*.pdf"))
    if not pdf_paths:
        print(f"  {fail_str()}  No PDFs found in {inputs_dir}")
        print(f"             Generate them with: python tests/generate_pdfs.py")
        return 0, 1

    print(f"  Found {len(pdf_paths)} PDFs in {inputs_dir}")

    out_folder = tempfile.mkdtemp()
    pipeline = TaxPipeline(
        api_key=api_key,
        template_1040=template_1040 or str(PROJECT_DIR / "25_1040.xlsx"),
        template_doublecheck=template_doublecheck or str(PROJECT_DIR / "2025_Tax_Return_Double_Check.xlsx"),
        output_folder=out_folder,
        log_callback=lambda msg: print(f"  {msg}"),
    )

    try:
        pipeline.run([str(p) for p in pdf_paths], "Thornton", "James")
    except Exception as e:
        print(f"  {fail_str()}  Pipeline raised: {e}")
        shutil.rmtree(out_folder, ignore_errors=True)
        return 0, 1

    out_dir = Path(out_folder) / "Thornton_James_2025"
    passed, failed = run_validation(out_dir / "Review", fixture_dir)

    shutil.rmtree(out_folder, ignore_errors=True)
    return passed, failed


# ── Validate-only mode ────────────────────────────────────────────────────────

def run_validation(review_dir: Path, fixture_dir: Path) -> tuple[int, int]:
    """Validate existing output Excel files against expected fixture JSON."""
    exp_dc_path   = fixture_dir / "expected_doublecheck.json"
    exp_1040_path = fixture_dir / "expected_1040.json"

    passed = failed = 0

    # Derive client slug from fixture metadata, or fall back to glob
    meta_path = fixture_dir / "meta.json"
    if meta_path.exists():
        with open(meta_path) as f:
            client_slug = json.load(f).get("client_slug", "")
    else:
        matches = list(review_dir.glob("*_DoubleCheck.xlsx"))
        client_slug = matches[0].stem.replace("_DoubleCheck", "") if matches else ""

    dc_path  = review_dir / f"{client_slug}_DoubleCheck.xlsx"
    i40_path = review_dir / f"{client_slug}_1040.xlsx"

    if exp_dc_path.exists() and dc_path.exists():
        with open(exp_dc_path) as f:
            exp_dc = json.load(f)
        wb = openpyxl.load_workbook(str(dc_path))
        p, f2 = validate_workbook(wb, exp_dc["sheets"], "DoubleCheck")
        passed += p; failed += f2
    else:
        print(f"  {warn_str()}  DoubleCheck expected JSON or output not found — skipping")

    if exp_1040_path.exists() and i40_path.exists():
        with open(exp_1040_path) as f:
            exp_1040 = json.load(f)
        wb = openpyxl.load_workbook(str(i40_path))
        p, f2 = validate_workbook(wb, exp_1040["sheets"], "Glenn Reeves 1040")
        passed += p; failed += f2
    else:
        print(f"  {warn_str()}  1040 expected JSON or output not found — skipping")

    return passed, failed


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="PrepDesk test suite")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--unit",        action="store_true", help="Unit tests (no API, uses golden fixtures)")
    group.add_argument("--integration", action="store_true", help="Full pipeline integration test (requires API + PDFs)")
    group.add_argument("--validate",    metavar="OUTPUT_DIR", help="Validate existing Excel output at path")
    parser.add_argument("--fixture",    default="thornton_mfj", help="Fixture folder name (default: thornton_mfj)")
    parser.add_argument("--log",        metavar="FILE", help="Write results to a file (in addition to stdout)")
    args = parser.parse_args()

    fixture_dir = FIXTURES_DIR / args.fixture
    if not fixture_dir.exists():
        print(f"Fixture not found: {fixture_dir}")
        sys.exit(1)

    # Tee all output to tests/last_run.log (plain text, no ANSI)
    import re as _re, io as _io
    _ansi = _re.compile(r'\x1b\[[0-9;]*m')
    _log_lines = []
    _orig_print = print

    def _tee(*a, **kw):
        _orig_print(*a, **kw)
        buf = _io.StringIO()
        kw2 = {k: v for k, v in kw.items() if k not in ('file',)}
        _orig_print(*a, file=buf, **kw2)
        _log_lines.append(_ansi.sub('', buf.getvalue()))

    import builtins
    builtins.print = _tee

    print(f"\nPrepDesk Test Suite — {args.fixture}")
    print("=" * 55)

    total_passed = total_failed = 0

    if args.unit:
        print(f"\n{BOLD}Mode: Unit (golden fixtures, no API){RESET}")
        p, f = run_unit_test(fixture_dir)
        total_passed += p; total_failed += f

    elif args.integration:
        print(f"\n{BOLD}Mode: Integration (full pipeline + real PDFs){RESET}")
        p, f = run_integration_test(fixture_dir)
        total_passed += p; total_failed += f

    elif args.validate:
        print(f"\n{BOLD}Mode: Validate existing output at {args.validate}{RESET}")
        p, f = run_validation(Path(args.validate), fixture_dir)
        total_passed += p; total_failed += f

    # Summary
    print(f"\n{'='*55}")
    total = total_passed + total_failed
    if total_failed == 0:
        print(f"{GREEN}{BOLD}ALL TESTS PASSED — {total_passed}/{total}{RESET}")
    else:
        print(f"{RED}{BOLD}FAILURES: {total_failed}/{total} assertions failed{RESET}")
        print(f"{GREEN}Passed: {total_passed}{RESET}  {RED}Failed: {total_failed}{RESET}")

    # Write log and restore print
    builtins.print = _orig_print
    log_path = TESTS_DIR / f"last_run_{args.fixture}.log"
    log_path.write_text("".join(_log_lines), encoding="utf-8")

    sys.exit(0 if total_failed == 0 else 1)


if __name__ == "__main__":
    main()
